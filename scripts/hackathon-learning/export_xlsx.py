#!/usr/bin/env python3
"""Export full hackathon learning analytics → multi-sheet xlsx.

Pulls from the DB pointed at by NEXT_PUBLIC_SUPABASE_URL using the service-role key
(bypasses RLS). Run against prod:

  pnpm exec dotenv -e .env.production -o -- python3 scripts/hackathon-learning/export_xlsx.py

Writes ./hackathon-learning-<date>.xlsx with sheets: Teams, Video, Cycles, Semifinal.
"""
import os, sys, json, datetime, urllib.request, urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font

URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
if not URL or not KEY:
    print("❌ missing NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY"); sys.exit(1)


def rest(table, select="*", **params):
    q = {"select": select, "limit": "5000", **params}
    req = urllib.request.Request(
        f"{URL}/rest/v1/{table}?{urllib.parse.urlencode(q)}",
        headers={"apikey": KEY, "Authorization": f"Bearer {KEY}"},
    )
    with urllib.request.urlopen(req) as r:
        return json.load(r)


def write_sheet(wb, title, rows, headers):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for i, h in enumerate(headers, 1):
        width = max(len(h), *(len(str(r.get(h, ""))) for r in rows)) if rows else len(h)
        ws.column_dimensions[ws.cell(1, i).column_letter].width = min(max(width + 2, 10), 70)


def main():
    teams = {t["id"]: t["name"] for t in rest("hackathon_teams", "id,name")}
    metrics = [m for m in rest("hackathon_learning_metrics") if m.get("subject_type") == "team"]
    signals = rest("hackathon_submission_signals")
    semis = rest("hackathon_semifinal_scores")

    videos = {s["team_id"]: (s.get("signals") or {}) for s in signals if s.get("submission_scope") == "phase4_video"}
    cycles = [s for s in signals if s.get("submission_scope") == "phase3_cycle"]

    # --- Teams sheet (one row per team, everything flat) ---
    team_rows = []
    for m in metrics:
        ev = m.get("evidence") or {}
        v = videos.get(m["team_id"], {})
        team_rows.append({
            "team": ev.get("label") or teams.get(m["team_id"], ""),
            "learning_index": m.get("learning_index"),
            "semifinal_total": m.get("semifinal_total"),
            "quadrant": m.get("quadrant"),
            "plan_fidelity": m.get("grappling_score"),
            "ai_likelihood": m.get("ai_likelihood"),
            "engagement": m.get("engagement_score"),
            "iteration": m.get("iteration_score"),
            "completion": ev.get("completion"),
            "members": ev.get("members"),
            "phase3_cycles": ev.get("phase3_cycles"),
            "scored_subs": ev.get("scored_subs"),
            "honest_wrongness": "yes" if ev.get("honest_wrongness") else "no",
            "video_total": v.get("total"),
            "video_ai_likelihood": v.get("ai_likelihood"),
            "journey_all_phases": m.get("journey_summary") or "",
            "journey_in_video": v.get("journey_summary") or "",
            "video_url": v.get("video_url") or "",
        })
    team_rows.sort(key=lambda r: (r["learning_index"] or 0), reverse=True)

    # --- Video sheet ---
    video_rows = []
    for tid, v in videos.items():
        video_rows.append({
            "team": teams.get(tid, ""),
            "total": v.get("total"), "ai_likelihood": v.get("ai_likelihood"),
            "story_clarity": v.get("story_clarity"), "evidence_integration": v.get("evidence_integration"),
            "delivery": v.get("delivery"), "demo_shown": "yes" if v.get("demo_shown") else "no",
            "journey": v.get("journey_summary") or "", "summary": v.get("summary") or "",
            "rationale": v.get("rationale") or "", "video_url": v.get("video_url") or "",
        })
    video_rows.sort(key=lambda r: (r["total"] or 0), reverse=True)

    # --- Cycles sheet ---
    cycle_rows = []
    for c in cycles:
        sj = c.get("signals") or {}
        cycle_rows.append({
            "team": teams.get(c["team_id"], ""),
            "total": sj.get("total"),
            "hypothesis_quality": sj.get("hypothesis_quality"),
            "variable_isolation": sj.get("variable_isolation"),
            "behavioral_evidence": sj.get("behavioral_evidence"),
            "tester_freshness": sj.get("tester_freshness"),
            "synthesis_honesty": sj.get("synthesis_honesty"),
            "honest_wrongness": "yes" if sj.get("honest_wrongness") else "no",
            "rationale": sj.get("rationale") or "",
        })
    cycle_rows.sort(key=lambda r: (r["team"], -(r["total"] or 0)))

    # --- Semifinal sheet ---
    semi_rows = []
    for s in semis:
        semi_rows.append({
            "team": teams.get(s.get("team_id"), s.get("team_id", "")),
            "division": s.get("division", ""), "panel": s.get("panel", ""),
            "total": s.get("total"), "judge_count": s.get("judge_count"),
        })
    semi_rows.sort(key=lambda r: (r["total"] or 0), reverse=True)

    wb = Workbook(); wb.remove(wb.active)
    write_sheet(wb, "Teams", team_rows, list(team_rows[0].keys()) if team_rows else ["team"])
    write_sheet(wb, "Video", video_rows, list(video_rows[0].keys()) if video_rows else ["team"])
    write_sheet(wb, "Cycles", cycle_rows, list(cycle_rows[0].keys()) if cycle_rows else ["team"])
    write_sheet(wb, "Semifinal", semi_rows, list(semi_rows[0].keys()) if semi_rows else ["team"])

    date = datetime.date.today().isoformat()
    out = f"hackathon-learning-{date}.xlsx"
    wb.save(out)
    print(f"✅ {out} — Teams={len(team_rows)} Video={len(video_rows)} Cycles={len(cycle_rows)} Semifinal={len(semi_rows)}")


if __name__ == "__main__":
    main()
